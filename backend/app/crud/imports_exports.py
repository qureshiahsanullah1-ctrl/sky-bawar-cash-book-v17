from __future__ import annotations

import re

from datetime import date, datetime, timedelta, timezone

from sqlalchemy import func, or_

from sqlalchemy.orm import Session

from .. import models, schemas
from .utils import _normalize_text, _amount, _date_value, _datetime_value, _first_value, _rows_from, _id_key, _map_imported, _get_imported, _choice, _backup_root, _normalize_settings_backup, _normalize_account_backup, _normalize_employee_backup, _transaction_type_from_backup, _normalize_transaction_backup, utcnow

from ..csv_import import parse_cashbook_csv
from .transactions import _cash_to_afn, _afn_to_usd, _next_transaction_no


def import_cashbook_csv(
    db: Session, content: str, filename: str = "cashbook.csv"
) -> dict:
    rows = parse_cashbook_csv(content)
    imported = 0
    skipped = 0
    created_accounts = 0

    existing_signatures = {
        (
            tx.date,
            tx.account_name.strip().lower(),
            tx.detail.strip().lower(),
            tx.transaction_type,
            _amount(tx.cash_in_afn),
            _amount(tx.cash_out_afn),
            _amount(tx.usd_in),
            _amount(tx.usd_out),
            _amount(tx.exchange_rate),
            tx.note.strip().lower(),
        )
        for tx in db.query(models.Transaction).all()
    }
    accounts = {
        account.name.strip().lower(): account
        for account in db.query(models.Account).all()
    }

    try:
        for payload in rows:
            amount_afn = _cash_to_afn(payload)
            derived_usd = _afn_to_usd(amount_afn, payload.exchange_rate)
            cash_in_afn = amount_afn if payload.transaction_type == "cash_in" else 0
            cash_out_afn = amount_afn if payload.transaction_type == "cash_out" else 0
            usd_in = (
                _amount(payload.usd_in)
                if payload.transaction_type == "cash_in" and _amount(payload.usd_in)
                else derived_usd if payload.transaction_type == "cash_in" else 0
            )
            usd_out = (
                _amount(payload.usd_out)
                if payload.transaction_type == "cash_out" and _amount(payload.usd_out)
                else derived_usd if payload.transaction_type == "cash_out" else 0
            )
            signature = (
                payload.date,
                payload.account_name.strip().lower(),
                payload.detail.strip().lower(),
                payload.transaction_type,
                cash_in_afn,
                cash_out_afn,
                usd_in,
                usd_out,
                _amount(payload.exchange_rate),
                payload.note.strip().lower(),
            )
            if signature in existing_signatures:
                skipped += 1
                continue

            account_key = payload.account_name.strip().lower()
            account = accounts.get(account_key)
            if not account:
                account = models.Account(
                    name=_normalize_text(payload.account_name),
                    opening_balance_afn=0,
                    opening_balance_usd=0,
                )
                db.add(account)
                db.flush()
                accounts[account_key] = account
                created_accounts += 1

            transaction = models.Transaction(
                transaction_no=_next_transaction_no(db, payload.date),
                date=payload.date,
                account_id=account.id,
                account_name=account.name,
                detail=_normalize_text(payload.detail),
                transaction_type=payload.transaction_type,
                cash_in_afn=cash_in_afn,
                cash_out_afn=cash_out_afn,
                usd_in=usd_in,
                usd_out=usd_out,
                exchange_rate=_amount(payload.exchange_rate),
                converted_afn=amount_afn,
                payment_method=payload.payment_method,
                category=payload.category,
                note=_normalize_text(payload.note),
            )
            db.add(transaction)
            db.flush()
            existing_signatures.add(signature)
            imported += 1

        db.add(
            models.BackupLog(
                backup_name=_normalize_text(filename) or "cashbook.csv",
                backup_type="csv_import",
                note=f"Imported {imported} transactions, skipped {skipped} duplicates, created {created_accounts} accounts",
            )
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "ok": True,
        "imported_transactions": imported,
        "skipped_duplicates": skipped,
        "created_accounts": created_accounts,
    }


def import_master_excel(db: Session, file_bytes: bytes, filename: str) -> dict:
    import io
    import openpyxl

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except Exception as error:
        raise ValueError(f"Failed to parse Excel file: {error}") from error

    sheets_processed = len(wb.sheetnames)
    imported_transactions = 0
    created_accounts = 0
    today_str = utcnow().strftime("%Y-%m-%d")

    accounts = {acc.name.strip().lower(): acc for acc in db.query(models.Account).all()}
    existing_sigs = set()
    for row in db.query(
        models.Transaction.date,
        models.Transaction.account_name,
        models.Transaction.detail,
        models.Transaction.transaction_type,
        models.Transaction.cash_in_afn,
        models.Transaction.cash_out_afn,
        models.Transaction.usd_in,
        models.Transaction.usd_out,
    ).all():
        existing_sigs.add(
            (
                str(row[0] or ""),
                (row[1] or "").strip().lower(),
                (row[2] or "").strip().lower(),
                str(row[3] or ""),
                float(row[4] or 0),
                float(row[5] or 0),
                float(row[6] or 0),
                float(row[7] or 0),
            )
        )

    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                continue

            header_idx = -1
            col_map = {}
            for idx, r in enumerate(rows[:10]):
                r_str = [str(c or "").strip().lower() for c in r]
                joined = " ".join(r_str)
                if any(
                    k in joined
                    for k in [
                        "date",
                        "detail",
                        "description",
                        "received",
                        "paid",
                        "amount",
                        "in",
                        "out",
                        "debit",
                        "credit",
                    ]
                ):
                    header_idx = idx
                    for c_idx, val in enumerate(r_str):
                        if "date" in val:
                            col_map["date"] = c_idx
                        elif any(
                            k in val for k in ["account", "name", "customer", "vendor"]
                        ):
                            col_map["account"] = c_idx
                        elif any(
                            k in val
                            for k in ["detail", "description", "particular", "item"]
                        ):
                            col_map["detail"] = c_idx
                        elif any(
                            k in val for k in ["received", "cash in", "debit"]
                        ) or (val == "in"):
                            col_map["cash_in"] = c_idx
                        elif any(k in val for k in ["paid", "cash out", "credit"]) or (
                            val == "out"
                        ):
                            col_map["cash_out"] = c_idx
                        elif "usd" in val:
                            col_map["usd"] = c_idx
                        elif any(k in val for k in ["rate", "exchange"]):
                            col_map["rate"] = c_idx
                        elif any(k in val for k in ["note", "remark", "ref"]):
                            col_map["note"] = c_idx
                    break

            if header_idx == -1:
                # Fallback mapping
                col_map = {"date": 0, "detail": 1, "cash_in": 2, "cash_out": 3}
                header_idx = 0

            for r in rows[header_idx + 1 :]:
                if not r or not any(r):
                    continue

                raw_date = (
                    r[col_map["date"]]
                    if "date" in col_map and col_map["date"] < len(r)
                    else None
                )
                if isinstance(raw_date, datetime):
                    tx_date = raw_date.strftime("%Y-%m-%d")
                elif isinstance(raw_date, date):
                    tx_date = raw_date.strftime("%Y-%m-%d")
                elif isinstance(raw_date, str) and raw_date.strip():
                    tx_date = raw_date.strip()[:10]
                else:
                    tx_date = today_str

                acc_cell = (
                    r[col_map["account"]]
                    if "account" in col_map and col_map["account"] < len(r)
                    else None
                )
                acc_name = str(acc_cell).strip() if acc_cell else sheet_name.strip()
                if not acc_name or acc_name.lower() in ("sheet", "total", "summary"):
                    acc_name = sheet_name.strip()

                detail_cell = (
                    r[col_map["detail"]]
                    if "detail" in col_map and col_map["detail"] < len(r)
                    else None
                )
                detail = (
                    str(detail_cell).strip()
                    if detail_cell
                    else f"Entry from {sheet_name}"
                )

                cin = (
                    _amount(r[col_map["cash_in"]])
                    if "cash_in" in col_map and col_map["cash_in"] < len(r)
                    else 0.0
                )
                cout = (
                    _amount(r[col_map["cash_out"]])
                    if "cash_out" in col_map and col_map["cash_out"] < len(r)
                    else 0.0
                )
                usd_amt = (
                    _amount(r[col_map["usd"]])
                    if "usd" in col_map and col_map["usd"] < len(r)
                    else 0.0
                )
                rate = (
                    _amount(r[col_map["rate"]])
                    if "rate" in col_map and col_map["rate"] < len(r)
                    else 0.0
                )
                note_cell = (
                    r[col_map["note"]]
                    if "note" in col_map and col_map["note"] < len(r)
                    else None
                )
                note = str(note_cell).strip() if note_cell else ""

                if cin == 0.0 and cout == 0.0 and usd_amt == 0.0:
                    continue

                tx_type = (
                    "cash_in"
                    if (cin > 0 or (usd_amt > 0 and cout == 0))
                    else "cash_out"
                )
                cash_in_afn = cin if tx_type == "cash_in" else 0.0
                cash_out_afn = cout if tx_type == "cash_out" else 0.0
                usd_in = usd_amt if tx_type == "cash_in" else 0.0
                usd_out = usd_amt if tx_type == "cash_out" else 0.0

                sig = (
                    tx_date,
                    acc_name.lower(),
                    detail.lower(),
                    tx_type,
                    cash_in_afn,
                    cash_out_afn,
                    usd_in,
                    usd_out,
                )
                if sig in existing_sigs:
                    continue

                acc_key = acc_name.lower()
                account = accounts.get(acc_key)
                if not account:
                    account = models.Account(
                        name=_normalize_text(acc_name),
                        opening_balance_afn=0,
                        opening_balance_usd=0,
                    )
                    db.add(account)
                    db.flush()
                    accounts[acc_key] = account
                    created_accounts += 1

                tx = models.Transaction(
                    transaction_no=_next_transaction_no(db, tx_date),
                    date=tx_date,
                    account_id=account.id,
                    account_name=account.name,
                    detail=_normalize_text(detail),
                    transaction_type=tx_type,
                    cash_in_afn=cash_in_afn,
                    cash_out_afn=cash_out_afn,
                    usd_in=usd_in,
                    usd_out=usd_out,
                    exchange_rate=rate,
                    converted_afn=cash_in_afn or cash_out_afn,
                    payment_method="cash",
                    category="other",
                    note=_normalize_text(note),
                )
                db.add(tx)
                db.flush()
                existing_sigs.add(sig)
                imported_transactions += 1

        db.add(
            models.BackupLog(
                backup_name=_normalize_text(filename) or "master-excel.xlsx",
                backup_type="excel_import",
                note=f"Parsed {sheets_processed} sheets. Imported {imported_transactions} transactions, created {created_accounts} accounts",
            )
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "ok": True,
        "message": "Master ledger Excel data imported successfully.",
        "filename": filename,
        "sheets_processed": sheets_processed,
        "imported_transactions": imported_transactions,
        "created_accounts": created_accounts,
    }


