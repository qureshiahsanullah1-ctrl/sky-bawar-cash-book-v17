from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from . import models, schemas
from .csv_import import parse_cashbook_csv


def _normalize_text(value: str | None) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _amount(value) -> float:
    try:
        if isinstance(value, str):
            cleaned = re.sub(r"[^0-9.\-]", "", value.replace(",", ""))
            value = cleaned or 0
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace("Z", "+00:00")
        for parser in (date.fromisoformat, datetime.fromisoformat):
            try:
                parsed = parser(text)
                return parsed.date() if isinstance(parsed, datetime) else parsed
            except ValueError:
                continue
        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return value


def _datetime_value(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            parsed_date = _date_value(text)
            return (
                datetime.combine(parsed_date, datetime.min.time())
                if isinstance(parsed_date, date)
                else None
            )
    return value


def _first_value(data: dict, *keys, default=None):
    for key in keys:
        if key in data and data[key] not in (None, ""):
            return data[key]
    return default


def _rows_from(payload: dict, *keys) -> list:
    for key in keys:
        value = payload.get(key)
        if isinstance(value, list):
            return value
        if isinstance(value, dict):
            if isinstance(value.get("rows"), list):
                return value["rows"]
            if isinstance(value.get("data"), list):
                return value["data"]
            if value and all(isinstance(item, dict) for item in value.values()):
                return list(value.values())
    return []


def _id_key(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return str(value)


def _map_imported(mapping: dict, original_id, value) -> None:
    key = _id_key(original_id)
    if key is None:
        return
    mapping[key] = value
    mapping[str(key)] = value


def _get_imported(mapping: dict, original_id):
    key = _id_key(original_id)
    if key is None:
        return None
    return mapping.get(key) or mapping.get(str(key))


def _choice(
    value, allowed: set[str], default: str, aliases: dict[str, str] | None = None
) -> str:
    text = _normalize_text(str(value or "")).lower().replace("-", "_").replace(" ", "_")
    aliases = aliases or {}
    return aliases.get(text) or (text if text in allowed else default)


def _backup_root(payload: dict) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("Backup file must be a JSON object.")
    for key in ("payload", "backup", "data"):
        nested = payload.get(key)
        if isinstance(nested, dict) and any(
            marker in nested
            for marker in (
                "settings",
                "accounts",
                "employees",
                "transactions",
                "records",
            )
        ):
            payload = nested
            break
    if "transactions" not in payload:
        for key in ("records", "cashbook", "cash_book", "entries"):
            if key in payload:
                payload = {**payload, "transactions": payload[key]}
                break
    return payload


def _normalize_settings_backup(settings_data: dict) -> dict:
    if not isinstance(settings_data, dict):
        return {}
    date_format = _choice(
        _first_value(
            settings_data,
            "date_display_format",
            "dateFormat",
            "calendar_format",
            default="dual",
        ),
        {"persian", "gregorian", "dual"},
        "dual",
        {
            "english": "gregorian",
            "jalali": "persian",
            "shamsi": "persian",
            "both": "dual",
        },
    )
    default_exchange_rate = _first_value(
        settings_data,
        "default_exchange_rate",
        "defaultExchangeRate",
        "exchange_rate",
        "exchangeRate",
    )
    normalized = {
        "company_name": _first_value(
            settings_data, "company_name", "companyName", "name"
        ),
        "company_phone": _first_value(
            settings_data, "company_phone", "companyPhone", "phone"
        ),
        "company_email": _first_value(
            settings_data, "company_email", "companyEmail", "email"
        ),
        "company_website": _first_value(
            settings_data, "company_website", "companyWebsite", "website"
        ),
        "company_tax_number": _first_value(
            settings_data, "company_tax_number", "companyTaxNumber", "tax_number"
        ),
        "company_logo": _first_value(
            settings_data, "company_logo", "companyLogo", "logo"
        ),
        "company_address": _first_value(
            settings_data, "company_address", "companyAddress", "address"
        ),
        "company_license": _first_value(
            settings_data, "company_license", "companyLicense", "license"
        ),
        "default_exchange_rate": (
            _amount(default_exchange_rate)
            if default_exchange_rate not in (None, "")
            else None
        ),
        "default_currency": str(
            _first_value(
                settings_data,
                "default_currency",
                "defaultCurrency",
                "currency",
                default="AFN",
            )
            or "AFN"
        ).upper(),
        "theme": _first_value(settings_data, "theme", default="dark"),
        "language": _first_value(settings_data, "language", default="English"),
        "date_display_format": date_format,
        "print_footer_text": _first_value(
            settings_data, "print_footer_text", "printFooterText", "footer"
        ),
        "auto_logout_minutes": int(
            _amount(
                _first_value(
                    settings_data,
                    "auto_logout_minutes",
                    "autoLogoutMinutes",
                    default=30,
                )
            )
            or 30
        ),
    }
    return {key: value for key, value in normalized.items() if value is not None}


def _normalize_account_backup(account_data: dict) -> dict | None:
    if not isinstance(account_data, dict):
        return None
    name = _normalize_text(
        _first_value(
            account_data,
            "name",
            "account_name",
            "accountName",
            "customer",
            "customer_name",
            "person",
            "party",
            "full_name",
        )
    )
    if not name:
        return None
    return {
        "name": name,
        "account_type": _choice(
            _first_value(account_data, "account_type", "accountType", "type"),
            {"customer", "supplier", "worker", "factory", "expense", "other"},
            "other",
            {
                "client": "customer",
                "employee": "worker",
                "staff": "worker",
                "vendor": "supplier",
                "company": "supplier",
            },
        ),
        "phone": _normalize_text(
            _first_value(account_data, "phone", "mobile", "contact")
        ),
        "address": _normalize_text(_first_value(account_data, "address", "location")),
        "opening_balance_afn": _amount(
            _first_value(
                account_data,
                "opening_balance_afn",
                "openingBalanceAfn",
                "opening_balance",
                "balance_afn",
                "balance",
            )
        ),
        "opening_balance_usd": _amount(
            _first_value(
                account_data, "opening_balance_usd", "openingBalanceUsd", "balance_usd"
            )
        ),
        "note": _normalize_text(
            _first_value(account_data, "note", "notes", "description")
        ),
    }


def _normalize_employee_backup(employee_data: dict) -> dict | None:
    if not isinstance(employee_data, dict):
        return None
    full_name = _normalize_text(
        _first_value(
            employee_data,
            "full_name",
            "fullName",
            "employee_name",
            "employeeName",
            "name",
        )
    )
    if not full_name:
        return None
    joining_date = _date_value(
        _first_value(employee_data, "joining_date", "joiningDate", "date_joined")
    )
    employment_end_date = _date_value(
        _first_value(
            employee_data,
            "employment_end_date",
            "employmentEndDate",
            "end_date",
            "endDate",
            "termination_date",
            "terminationDate",
        )
    )
    return {
        "full_name": full_name,
        "father_name": _normalize_text(
            _first_value(employee_data, "father_name", "fatherName", "father")
        ),
        "phone": _normalize_text(
            _first_value(employee_data, "phone", "mobile", "contact")
        ),
        "position": _normalize_text(
            _first_value(
                employee_data,
                "position",
                "role",
                "job_title",
                "jobTitle",
                "designation",
            )
        )
        or "Employee",
        "department": _normalize_text(
            _first_value(employee_data, "department", "section")
        ),
        "joining_date": joining_date,
        "employment_end_date": employment_end_date,
        "monthly_salary": _amount(
            _first_value(
                employee_data, "monthly_salary", "monthlySalary", "salary", "salary_afn"
            )
        ),
        "currency": (
            "USD"
            if str(_first_value(employee_data, "currency", default="AFN")).upper()
            == "USD"
            else "AFN"
        ),
        "avatar_url": _normalize_text(
            _first_value(
                employee_data,
                "avatar_url",
                "avatarUrl",
                "avatar",
                "avatar_path",
                "avatarPath",
            )
        ),
        "status": _choice(
            _first_value(employee_data, "status", "is_active", "active"),
            {"active", "inactive"},
            "active",
            {"true": "active", "false": "inactive", "1": "active", "0": "inactive"},
        ),
        "notes": _normalize_text(_first_value(employee_data, "notes", "note")),
    }


def _transaction_type_from_backup(
    tx_data: dict,
    cash_in_afn: float,
    cash_out_afn: float,
    usd_in: float,
    usd_out: float,
) -> str:
    tx_type = _choice(
        _first_value(
            tx_data, "transaction_type", "transactionType", "type", "kind", "direction"
        ),
        {"cash_in", "cash_out"},
        "",
        {
            "cashin": "cash_in",
            "in": "cash_in",
            "income": "cash_in",
            "credit": "cash_in",
            "receive": "cash_in",
            "received": "cash_in",
            "cashout": "cash_out",
            "out": "cash_out",
            "expense": "cash_out",
            "debit": "cash_out",
            "payment": "cash_out",
            "paid": "cash_out",
        },
    )
    if tx_type:
        return tx_type
    amount = _amount(_first_value(tx_data, "amount", "total"))
    if cash_in_afn or usd_in or amount > 0:
        return "cash_in"
    if cash_out_afn or usd_out or amount < 0:
        return "cash_out"
    return "cash_out"


def _normalize_transaction_backup(tx_data: dict) -> dict | None:
    if not isinstance(tx_data, dict):
        return None
    tx_date = _date_value(
        _first_value(
            tx_data,
            "date",
            "transaction_date",
            "transactionDate",
            "created_at",
            "createdAt",
        )
    )
    account_name = (
        _normalize_text(
            _first_value(
                tx_data,
                "account_name",
                "accountName",
                "account",
                "name",
                "customer",
                "customer_name",
                "person",
                "party",
                "employee_name",
            )
        )
        or "Imported Account"
    )
    cash_in_afn = _amount(
        _first_value(
            tx_data, "cash_in_afn", "cashInAfn", "cash_in", "cashIn", "afn_in", "afnIn"
        )
    )
    cash_out_afn = _amount(
        _first_value(
            tx_data,
            "cash_out_afn",
            "cashOutAfn",
            "cash_out",
            "cashOut",
            "afn_out",
            "afnOut",
        )
    )
    usd_in = _amount(_first_value(tx_data, "usd_in", "usdIn", "dollar_in", "dollarIn"))
    usd_out = _amount(
        _first_value(tx_data, "usd_out", "usdOut", "dollar_out", "dollarOut")
    )
    tx_type = _transaction_type_from_backup(
        tx_data, cash_in_afn, cash_out_afn, usd_in, usd_out
    )
    amount = _amount(_first_value(tx_data, "amount", "total"))
    if amount and not any((cash_in_afn, cash_out_afn, usd_in, usd_out)):
        if tx_type == "cash_in":
            cash_in_afn = abs(amount)
        else:
            cash_out_afn = abs(amount)
    exchange_rate = _amount(
        _first_value(tx_data, "exchange_rate", "exchangeRate", "rate")
    )
    if (usd_in or usd_out) and not exchange_rate:
        exchange_rate = 64.3
    return {
        "id": _id_key(_first_value(tx_data, "id", "transaction_id", "transactionId")),
        "date": tx_date or date.today(),
        "account_id": _id_key(_first_value(tx_data, "account_id", "accountId")),
        "employee_id": _id_key(_first_value(tx_data, "employee_id", "employeeId")),
        "salary_month": _date_value(
            _first_value(tx_data, "salary_month", "salaryMonth", "month")
        ),
        "payroll_kind": _choice(
            _first_value(tx_data, "payroll_kind", "payrollKind"),
            {"salary", "advance"},
            None,
        ),
        "account_name": account_name,
        "detail": _normalize_text(
            _first_value(
                tx_data,
                "detail",
                "description",
                "memo",
                "particulars",
                "reason",
                "note",
            )
        )
        or "Imported transaction",
        "transaction_type": tx_type,
        "cash_in_afn": cash_in_afn,
        "cash_out_afn": cash_out_afn,
        "usd_in": usd_in,
        "usd_out": usd_out,
        "exchange_rate": exchange_rate,
        "converted_afn": _amount(
            _first_value(tx_data, "converted_afn", "convertedAfn", "afn_amount")
        ),
        "payment_method": _choice(
            _first_value(tx_data, "payment_method", "paymentMethod", "method"),
            {"cash", "bank", "hawala", "other"},
            "cash",
            {"transfer": "bank", "card": "bank", "cheque": "bank", "check": "bank"},
        ),
        "category": _choice(
            _first_value(tx_data, "category", "expense_category", "expenseCategory"),
            {
                "salary",
                "rent",
                "factory_expense",
                "home_expense",
                "bottles_account",
                "office_expense",
                "other",
            },
            "other",
            {
                "factory": "factory_expense",
                "home": "home_expense",
                "office": "office_expense",
                "bottle": "bottles_account",
                "bottles": "bottles_account",
                "employee_salary": "salary",
            },
        ),
        "note": _normalize_text(_first_value(tx_data, "note", "notes", "remarks")),
    }


def _ensure_settings(db: Session) -> models.Setting:
    settings = db.query(models.Setting).first()
    if settings:
        return settings
    settings = models.Setting()
    db.add(settings)
    db.commit()
    db.refresh(settings)
    return settings


def get_settings(db: Session) -> models.Setting:
    return _ensure_settings(db)


def update_settings(db: Session, payload: schemas.SettingUpdate) -> models.Setting:
    settings = _ensure_settings(db)
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(settings, key, value)
    db.commit()
    db.refresh(settings)
    return settings


def list_accounts(db: Session) -> list[models.Account]:
    return db.query(models.Account).order_by(models.Account.name.asc()).all()


def get_account(db: Session, account_id: int) -> models.Account | None:
    return db.query(models.Account).filter(models.Account.id == account_id).first()


def get_account_by_name(db: Session, name: str) -> models.Account | None:
    return (
        db.query(models.Account)
        .filter(func.lower(models.Account.name) == name.lower())
        .first()
    )


def create_account(db: Session, payload: schemas.AccountCreate) -> models.Account:
    account = get_account_by_name(db, payload.name)
    if account:
        raise ValueError("An account with this name already exists")
    account = models.Account(
        name=_normalize_text(payload.name),
        account_type=payload.account_type,
        phone=_normalize_text(payload.phone),
        address=_normalize_text(payload.address),
        opening_balance_afn=_amount(payload.opening_balance_afn),
        opening_balance_usd=_amount(payload.opening_balance_usd),
        note=_normalize_text(payload.note),
    )
    db.add(account)
    db.commit()
    db.refresh(account)
    return account


def update_account(
    db: Session, account: models.Account, payload: schemas.AccountUpdate
) -> models.Account:
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        if value is None:
            continue
        if key == "account_type":
            setattr(account, key, value)
        else:
            setattr(
                account,
                key,
                (
                    _amount(value)
                    if key.startswith("opening_balance")
                    else _normalize_text(value)
                ),
            )
    db.commit()
    db.refresh(account)
    return account


def delete_account(db: Session, account: models.Account) -> None:
    db.delete(account)
    db.commit()


def list_transactions(db: Session) -> list[models.Transaction]:
    return (
        db.query(models.Transaction)
        .order_by(models.Transaction.date.asc(), models.Transaction.id.asc())
        .all()
    )


def get_transaction(
    db: Session, transaction_id: int, user: models.User | None = None
) -> models.Transaction | None:
    query = db.query(models.Transaction).filter(models.Transaction.id == transaction_id)
    if user and user.role not in ["Administrator", "Super Admin"]:
        if user.assigned_branch_id is not None:
            query = query.filter(
                models.Transaction.branch_id == user.assigned_branch_id
            )
        elif user.assigned_group_id is not None:
            group_branch_ids = [
                b.id
                for b in db.query(models.Branch)
                .filter(models.Branch.group_id == user.assigned_group_id)
                .all()
            ]
            query = query.filter(models.Transaction.branch_id.in_(group_branch_ids))
    return query.first()


def _cash_to_afn(payload: schemas.TransactionBase) -> float:
    if payload.cash_in_afn or payload.cash_out_afn:
        return _amount(payload.cash_in_afn or payload.cash_out_afn)
    if payload.usd_in or payload.usd_out:
        usd_amount = _amount(payload.usd_in or payload.usd_out)
        return round(usd_amount * _amount(payload.exchange_rate), 2)
    return 0.0


def _afn_to_usd(amount_afn: float, rate: float) -> float:
    return (
        round(_amount(amount_afn) / _amount(rate), 2)
        if _amount(amount_afn) and _amount(rate)
        else 0.0
    )


def _next_transaction_no(db: Session, transaction_date: date) -> str:
    if isinstance(transaction_date, str):
        try:
            transaction_date = datetime.strptime(transaction_date.replace("/", "-"), "%Y-%m-%d").date()
        except Exception:
            transaction_date = date.today()
    prefix = transaction_date.strftime("TX-%Y%m%d")
    existing = (
        db.query(models.Transaction.transaction_no)
        .filter(models.Transaction.transaction_no.like(f"{prefix}-%"))
        .all()
    )
    sequences = []
    for (number,) in existing:
        try:
            sequences.append(int(str(number).rsplit("-", 1)[-1]))
        except (TypeError, ValueError):
            continue
    return f"{prefix}-{(max(sequences, default=0) + 1):04d}"


def _validate_transaction(payload: schemas.TransactionBase) -> None:
    values = [
        payload.cash_in_afn,
        payload.cash_out_afn,
        payload.usd_in,
        payload.usd_out,
        payload.exchange_rate,
    ]
    if any(_amount(value) < 0 for value in values):
        raise ValueError("Amounts cannot be negative")
    usd_amount = _amount(payload.usd_in or payload.usd_out)
    afn_amount = _amount(payload.cash_in_afn or payload.cash_out_afn)
    if afn_amount <= 0 and usd_amount <= 0:
        raise ValueError("At least one amount is required")
    if usd_amount > 0 and _amount(payload.exchange_rate) <= 0:
        raise ValueError("Exchange rate is required when USD is entered")


def create_transaction(
    db: Session, payload: schemas.TransactionCreate
) -> models.Transaction:
    _validate_transaction(payload)
    amount_afn = _cash_to_afn(payload)
    derived_usd = _afn_to_usd(amount_afn, payload.exchange_rate)
    employee = (
        db.query(models.Employee)
        .filter(models.Employee.id == payload.employee_id)
        .first()
        if payload.employee_id
        else None
    )
    if payload.employee_id and not employee:
        raise ValueError("Employee not found")
    if employee and payload.transaction_type != "cash_out":
        raise ValueError("Employee salary payments must be Cash Out")
    account = get_account(db, payload.account_id) if payload.account_id else None
    if employee:
        account = employee.account
    target_name = (_normalize_text(payload.account_name) or _normalize_text(payload.detail) or "General Account")[:250]
    if not account:
        account = get_account_by_name(db, target_name)
    if not account:
        account = models.Account(
            name=target_name,
            opening_balance_afn=0,
            opening_balance_usd=0,
            company_id=(getattr(payload, "company_id", None) or "bawar-star")
        )
        db.add(account)
        try:
            db.commit()
            db.refresh(account)
        except Exception:
            db.rollback()
            account = get_account_by_name(db, target_name) or db.query(models.Account).first()
            if not account:
                account = models.Account(
                    name="General Account",
                    opening_balance_afn=0,
                    opening_balance_usd=0,
                    company_id="bawar-star"
                )
                db.add(account)
                try:
                    db.commit()
                    db.refresh(account)
                except Exception:
                    db.rollback()
                    account = db.query(models.Account).first()
    if not account:
        raise ValueError("Target account is required")

    detail_text = _normalize_text(payload.detail) or _normalize_text(payload.account_name) or ("Cash In" if payload.transaction_type == "cash_in" else "Cash Out")

    last_error = None
    for attempt in range(5):
        try:
            tx_no = _next_transaction_no(db, payload.date)
            transaction = models.Transaction(
                transaction_no=tx_no,
                date=payload.date,
                account_id=account.id,
                employee_id=employee.id if employee else None,
                company_id=(
                    (employee.company_id or "all")
                    if employee
                    else (getattr(payload, "company_id", None) or "bawar-star")
                ),
                salary_month=(
                    (payload.salary_month or payload.date).replace(day=1) if employee else None
                ),
                payroll_kind=(payload.payroll_kind or "salary") if employee else None,
                account_name=account.name,
                detail=detail_text,
                transaction_type=payload.transaction_type,
                cash_in_afn=amount_afn if payload.transaction_type == "cash_in" else 0,
                cash_out_afn=amount_afn if payload.transaction_type == "cash_out" else 0,
                usd_in=(
                    _amount(payload.usd_in)
                    if payload.transaction_type == "cash_in" and _amount(payload.usd_in)
                    else (derived_usd if payload.transaction_type == "cash_in" else 0)
                ),
                usd_out=(
                    _amount(payload.usd_out)
                    if payload.transaction_type == "cash_out" and _amount(payload.usd_out)
                    else (derived_usd if payload.transaction_type == "cash_out" else 0)
                ),
                exchange_rate=_amount(payload.exchange_rate),
                converted_afn=amount_afn,
                payment_method=payload.payment_method,
                category="salary" if employee else payload.category,
                note=_normalize_text(payload.note),
                branch_id=payload.branch_id,
            )
            db.add(transaction)
            db.commit()
            db.refresh(transaction)
            return transaction
        except Exception as err:
            db.rollback()
            last_error = err
    
    raise ValueError(f"Failed to record transaction due to database collision: {str(last_error)}")


def import_cashbook_csv(
    db: Session, content: str, filename: str = "cashbook.csv"
) -> dict:
    rows = parse_cashbook_csv(content)
    imported = 0
    skipped = 0
    created_accounts = 0

    existing_signatures = {
        (
            tx.date,
            tx.account_name.strip().lower(),
            tx.detail.strip().lower(),
            tx.transaction_type,
            _amount(tx.cash_in_afn),
            _amount(tx.cash_out_afn),
            _amount(tx.usd_in),
            _amount(tx.usd_out),
            _amount(tx.exchange_rate),
            tx.note.strip().lower(),
        )
        for tx in db.query(models.Transaction).all()
    }
    accounts = {
        account.name.strip().lower(): account
        for account in db.query(models.Account).all()
    }

    try:
        for payload in rows:
            amount_afn = _cash_to_afn(payload)
            derived_usd = _afn_to_usd(amount_afn, payload.exchange_rate)
            cash_in_afn = amount_afn if payload.transaction_type == "cash_in" else 0
            cash_out_afn = amount_afn if payload.transaction_type == "cash_out" else 0
            usd_in = (
                _amount(payload.usd_in)
                if payload.transaction_type == "cash_in" and _amount(payload.usd_in)
                else derived_usd if payload.transaction_type == "cash_in" else 0
            )
            usd_out = (
                _amount(payload.usd_out)
                if payload.transaction_type == "cash_out" and _amount(payload.usd_out)
                else derived_usd if payload.transaction_type == "cash_out" else 0
            )
            signature = (
                payload.date,
                payload.account_name.strip().lower(),
                payload.detail.strip().lower(),
                payload.transaction_type,
                cash_in_afn,
                cash_out_afn,
                usd_in,
                usd_out,
                _amount(payload.exchange_rate),
                payload.note.strip().lower(),
            )
            if signature in existing_signatures:
                skipped += 1
                continue

            account_key = payload.account_name.strip().lower()
            account = accounts.get(account_key)
            if not account:
                account = models.Account(
                    name=_normalize_text(payload.account_name),
                    opening_balance_afn=0,
                    opening_balance_usd=0,
                )
                db.add(account)
                db.flush()
                accounts[account_key] = account
                created_accounts += 1

            transaction = models.Transaction(
                transaction_no=_next_transaction_no(db, payload.date),
                date=payload.date,
                account_id=account.id,
                account_name=account.name,
                detail=_normalize_text(payload.detail),
                transaction_type=payload.transaction_type,
                cash_in_afn=cash_in_afn,
                cash_out_afn=cash_out_afn,
                usd_in=usd_in,
                usd_out=usd_out,
                exchange_rate=_amount(payload.exchange_rate),
                converted_afn=amount_afn,
                payment_method=payload.payment_method,
                category=payload.category,
                note=_normalize_text(payload.note),
            )
            db.add(transaction)
            db.flush()
            existing_signatures.add(signature)
            imported += 1

        db.add(
            models.BackupLog(
                backup_name=_normalize_text(filename) or "cashbook.csv",
                backup_type="csv_import",
                note=f"Imported {imported} transactions, skipped {skipped} duplicates, created {created_accounts} accounts",
            )
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "ok": True,
        "imported_transactions": imported,
        "skipped_duplicates": skipped,
        "created_accounts": created_accounts,
    }


def import_master_excel(db: Session, file_bytes: bytes, filename: str) -> dict:
    import io
    import openpyxl

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except Exception as error:
        raise ValueError(f"Failed to parse Excel file: {error}") from error

    sheets_processed = len(wb.sheetnames)
    imported_transactions = 0
    created_accounts = 0
    today_str = datetime.utcnow().strftime("%Y-%m-%d")

    accounts = {acc.name.strip().lower(): acc for acc in db.query(models.Account).all()}
    existing_sigs = set()
    for row in db.query(
        models.Transaction.date,
        models.Transaction.account_name,
        models.Transaction.detail,
        models.Transaction.transaction_type,
        models.Transaction.cash_in_afn,
        models.Transaction.cash_out_afn,
        models.Transaction.usd_in,
        models.Transaction.usd_out,
    ).all():
        existing_sigs.add(
            (
                str(row[0] or ""),
                (row[1] or "").strip().lower(),
                (row[2] or "").strip().lower(),
                str(row[3] or ""),
                float(row[4] or 0),
                float(row[5] or 0),
                float(row[6] or 0),
                float(row[7] or 0),
            )
        )

    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                continue

            header_idx = -1
            col_map = {}
            for idx, r in enumerate(rows[:10]):
                r_str = [str(c or "").strip().lower() for c in r]
                joined = " ".join(r_str)
                if any(
                    k in joined
                    for k in [
                        "date",
                        "detail",
                        "description",
                        "received",
                        "paid",
                        "amount",
                        "in",
                        "out",
                        "debit",
                        "credit",
                    ]
                ):
                    header_idx = idx
                    for c_idx, val in enumerate(r_str):
                        if "date" in val:
                            col_map["date"] = c_idx
                        elif any(
                            k in val for k in ["account", "name", "customer", "vendor"]
                        ):
                            col_map["account"] = c_idx
                        elif any(
                            k in val
                            for k in ["detail", "description", "particular", "item"]
                        ):
                            col_map["detail"] = c_idx
                        elif any(
                            k in val for k in ["received", "cash in", "debit"]
                        ) or (val == "in"):
                            col_map["cash_in"] = c_idx
                        elif any(k in val for k in ["paid", "cash out", "credit"]) or (
                            val == "out"
                        ):
                            col_map["cash_out"] = c_idx
                        elif "usd" in val:
                            col_map["usd"] = c_idx
                        elif any(k in val for k in ["rate", "exchange"]):
                            col_map["rate"] = c_idx
                        elif any(k in val for k in ["note", "remark", "ref"]):
                            col_map["note"] = c_idx
                    break

            if header_idx == -1:
                # Fallback mapping
                col_map = {"date": 0, "detail": 1, "cash_in": 2, "cash_out": 3}
                header_idx = 0

            for r in rows[header_idx + 1 :]:
                if not r or not any(r):
                    continue

                raw_date = (
                    r[col_map["date"]]
                    if "date" in col_map and col_map["date"] < len(r)
                    else None
                )
                if isinstance(raw_date, datetime):
                    tx_date = raw_date.strftime("%Y-%m-%d")
                elif isinstance(raw_date, date):
                    tx_date = raw_date.strftime("%Y-%m-%d")
                elif isinstance(raw_date, str) and raw_date.strip():
                    tx_date = raw_date.strip()[:10]
                else:
                    tx_date = today_str

                acc_cell = (
                    r[col_map["account"]]
                    if "account" in col_map and col_map["account"] < len(r)
                    else None
                )
                acc_name = str(acc_cell).strip() if acc_cell else sheet_name.strip()
                if not acc_name or acc_name.lower() in ("sheet", "total", "summary"):
                    acc_name = sheet_name.strip()

                detail_cell = (
                    r[col_map["detail"]]
                    if "detail" in col_map and col_map["detail"] < len(r)
                    else None
                )
                detail = (
                    str(detail_cell).strip()
                    if detail_cell
                    else f"Entry from {sheet_name}"
                )

                cin = (
                    _amount(r[col_map["cash_in"]])
                    if "cash_in" in col_map and col_map["cash_in"] < len(r)
                    else 0.0
                )
                cout = (
                    _amount(r[col_map["cash_out"]])
                    if "cash_out" in col_map and col_map["cash_out"] < len(r)
                    else 0.0
                )
                usd_amt = (
                    _amount(r[col_map["usd"]])
                    if "usd" in col_map and col_map["usd"] < len(r)
                    else 0.0
                )
                rate = (
                    _amount(r[col_map["rate"]])
                    if "rate" in col_map and col_map["rate"] < len(r)
                    else 0.0
                )
                note_cell = (
                    r[col_map["note"]]
                    if "note" in col_map and col_map["note"] < len(r)
                    else None
                )
                note = str(note_cell).strip() if note_cell else ""

                if cin == 0.0 and cout == 0.0 and usd_amt == 0.0:
                    continue

                tx_type = (
                    "cash_in"
                    if (cin > 0 or (usd_amt > 0 and cout == 0))
                    else "cash_out"
                )
                cash_in_afn = cin if tx_type == "cash_in" else 0.0
                cash_out_afn = cout if tx_type == "cash_out" else 0.0
                usd_in = usd_amt if tx_type == "cash_in" else 0.0
                usd_out = usd_amt if tx_type == "cash_out" else 0.0

                sig = (
                    tx_date,
                    acc_name.lower(),
                    detail.lower(),
                    tx_type,
                    cash_in_afn,
                    cash_out_afn,
                    usd_in,
                    usd_out,
                )
                if sig in existing_sigs:
                    continue

                acc_key = acc_name.lower()
                account = accounts.get(acc_key)
                if not account:
                    account = models.Account(
                        name=_normalize_text(acc_name),
                        opening_balance_afn=0,
                        opening_balance_usd=0,
                    )
                    db.add(account)
                    db.flush()
                    accounts[acc_key] = account
                    created_accounts += 1

                tx = models.Transaction(
                    transaction_no=_next_transaction_no(db, tx_date),
                    date=tx_date,
                    account_id=account.id,
                    account_name=account.name,
                    detail=_normalize_text(detail),
                    transaction_type=tx_type,
                    cash_in_afn=cash_in_afn,
                    cash_out_afn=cash_out_afn,
                    usd_in=usd_in,
                    usd_out=usd_out,
                    exchange_rate=rate,
                    converted_afn=cash_in_afn or cash_out_afn,
                    payment_method="cash",
                    category="other",
                    note=_normalize_text(note),
                )
                db.add(tx)
                db.flush()
                existing_sigs.add(sig)
                imported_transactions += 1

        db.add(
            models.BackupLog(
                backup_name=_normalize_text(filename) or "master-excel.xlsx",
                backup_type="excel_import",
                note=f"Parsed {sheets_processed} sheets. Imported {imported_transactions} transactions, created {created_accounts} accounts",
            )
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "ok": True,
        "message": "Master ledger Excel data imported successfully.",
        "filename": filename,
        "sheets_processed": sheets_processed,
        "imported_transactions": imported_transactions,
        "created_accounts": created_accounts,
    }


def update_transaction(
    db: Session, transaction: models.Transaction, payload: schemas.TransactionUpdate
) -> models.Transaction:
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        if value is None:
            continue
        if key in {"account_name", "detail", "note"}:
            setattr(transaction, key, _normalize_text(value))
        elif key in {
            "cash_in_afn",
            "cash_out_afn",
            "usd_in",
            "usd_out",
            "exchange_rate",
            "converted_afn",
        }:
            setattr(transaction, key, _amount(value))
        else:
            setattr(transaction, key, value)
    if transaction.transaction_type == "cash_in":
        transaction.cash_out_afn = 0
        transaction.usd_out = 0
        if not transaction.cash_in_afn:
            transaction.cash_in_afn = (
                round(
                    _amount(transaction.usd_in) * _amount(transaction.exchange_rate), 2
                )
                if transaction.usd_in and transaction.exchange_rate
                else transaction.cash_in_afn
            )
    if transaction.transaction_type == "cash_out":
        transaction.cash_in_afn = 0
        transaction.usd_in = 0
        if not transaction.cash_out_afn:
            transaction.cash_out_afn = (
                round(
                    _amount(transaction.usd_out) * _amount(transaction.exchange_rate), 2
                )
                if transaction.usd_out and transaction.exchange_rate
                else transaction.cash_out_afn
            )
    transaction.converted_afn = _amount(
        transaction.cash_in_afn or transaction.cash_out_afn
    )
    if not transaction.converted_afn and transaction.exchange_rate:
        transaction.converted_afn = round(
            _amount(transaction.usd_in or transaction.usd_out)
            * _amount(transaction.exchange_rate),
            2,
        )
    if transaction.exchange_rate:
        if (
            transaction.transaction_type == "cash_in"
            and transaction.cash_in_afn
            and not transaction.usd_in
        ):
            transaction.usd_in = _afn_to_usd(
                transaction.cash_in_afn, transaction.exchange_rate
            )
        if (
            transaction.transaction_type == "cash_out"
            and transaction.cash_out_afn
            and not transaction.usd_out
        ):
            transaction.usd_out = _afn_to_usd(
                transaction.cash_out_afn, transaction.exchange_rate
            )
    if transaction.account_name:
        account = (
            get_account(db, transaction.account_id) if transaction.account_id else None
        )
        if not account or account.name.lower() != transaction.account_name.lower():
            account = get_account_by_name(db, transaction.account_name)
            if not account:
                account = models.Account(
                    name=_normalize_text(transaction.account_name),
                    opening_balance_afn=0,
                    opening_balance_usd=0,
                )
                db.add(account)
                db.flush()
            transaction.account_id = account.id
            transaction.account_name = account.name
    _validate_transaction(
        schemas.TransactionCreate(
            date=transaction.date,
            account_id=transaction.account_id,
            account_name=transaction.account_name,
            detail=transaction.detail,
            transaction_type=transaction.transaction_type,
            cash_in_afn=transaction.cash_in_afn,
            cash_out_afn=transaction.cash_out_afn,
            usd_in=transaction.usd_in,
            usd_out=transaction.usd_out,
            exchange_rate=transaction.exchange_rate,
            converted_afn=transaction.converted_afn,
            payment_method=transaction.payment_method,
            category=transaction.category,
            note=transaction.note,
        )
    )
    db.commit()
    db.refresh(transaction)
    return transaction


def delete_transaction(db: Session, transaction: models.Transaction) -> None:
    db.delete(transaction)
    db.commit()


def summary(
    db: Session,
    user: models.User | None = None,
    group_id: int | None = None,
    branch_id: int | None = None,
) -> dict:
    today = date.today()
    month_start = today.replace(day=1)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)

    base_q = db.query(models.Transaction)
    if branch_id:
        base_q = base_q.filter(models.Transaction.branch_id == branch_id)
    elif group_id:
        branch_ids = [
            b.id
            for b in db.query(models.Branch)
            .filter(models.Branch.group_id == group_id)
            .all()
        ]
        base_q = base_q.filter(models.Transaction.branch_id.in_(branch_ids))

    if user and user.role not in ["Administrator", "Super Admin"]:
        if user.assigned_branch_id is not None:
            base_q = base_q.filter(
                models.Transaction.branch_id == user.assigned_branch_id
            )
        elif user.assigned_group_id is not None:
            group_branch_ids = [
                b.id
                for b in db.query(models.Branch)
                .filter(models.Branch.group_id == user.assigned_group_id)
                .all()
            ]
            base_q = base_q.filter(models.Transaction.branch_id.in_(group_branch_ids))

    cash_in_afn = _amount(
        base_q.with_entities(func.sum(models.Transaction.cash_in_afn)).scalar()
    )
    cash_out_afn = _amount(
        base_q.with_entities(func.sum(models.Transaction.cash_out_afn)).scalar()
    )
    usd_in = _amount(base_q.with_entities(func.sum(models.Transaction.usd_in)).scalar())
    usd_out = _amount(
        base_q.with_entities(func.sum(models.Transaction.usd_out)).scalar()
    )
    today_q = base_q.filter(models.Transaction.date == today)
    month_q = base_q.filter(
        models.Transaction.date >= month_start,
        models.Transaction.date < next_month,
    )

    today_transactions = today_q.count()
    monthly_transactions = month_q.count()

    today_cash_in = _amount(today_q.with_entities(func.sum(models.Transaction.cash_in_afn)).scalar())
    today_cash_out = _amount(today_q.with_entities(func.sum(models.Transaction.cash_out_afn)).scalar())
    monthly_cash_in = _amount(month_q.with_entities(func.sum(models.Transaction.cash_in_afn)).scalar())
    monthly_cash_out = _amount(month_q.with_entities(func.sum(models.Transaction.cash_out_afn)).scalar())

    return {
        "cash_in_afn": cash_in_afn,
        "cash_out_afn": cash_out_afn,
        "afn_balance": round(cash_in_afn - cash_out_afn, 2),
        "usd_in": usd_in,
        "usd_out": usd_out,
        "usd_balance": round(usd_in - usd_out, 2),
        "today_transactions": today_transactions,
        "monthly_transactions": monthly_transactions,
        "today_cash_in": today_cash_in,
        "today_cash_out": today_cash_out,
        "monthly_cash_in": monthly_cash_in,
        "monthly_cash_out": monthly_cash_out,
    }


def filtered_transactions(
    db: Session,
    user: models.User | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    type: str | None = None,
    search: str | None = None,
    account: str | None = None,
    category: str | None = None,
    payment_method: str | None = None,
    group_id: int | None = None,
    branch_id: int | None = None,
    skip: int = 0,
    limit: int | None = None,
) -> list[models.Transaction]:
    query = db.query(models.Transaction)
    if user and user.role not in ["Administrator", "Super Admin"]:
        if user.assigned_branch_id is not None:
            query = query.filter(
                models.Transaction.branch_id == user.assigned_branch_id
            )
        elif user.assigned_group_id is not None:
            group_branch_ids = [
                b.id
                for b in db.query(models.Branch)
                .filter(models.Branch.group_id == user.assigned_group_id)
                .all()
            ]
            query = query.filter(models.Transaction.branch_id.in_(group_branch_ids))

    if start_date:
        query = query.filter(models.Transaction.date >= start_date)
    if end_date:
        query = query.filter(models.Transaction.date <= end_date)
    if type in {"cash_in", "cash_out"}:
        query = query.filter(models.Transaction.transaction_type == type)
    if account:
        query = query.filter(
            func.lower(models.Transaction.account_name).like(f"%{account.lower()}%")
        )
    if category:
        query = query.filter(models.Transaction.category == category)
    if payment_method:
        query = query.filter(models.Transaction.payment_method == payment_method)
    if branch_id:
        query = query.filter(models.Transaction.branch_id == branch_id)
    elif group_id:
        branch_ids = [
            b.id
            for b in db.query(models.Branch)
            .filter(models.Branch.group_id == group_id)
            .all()
        ]
        query = query.filter(models.Transaction.branch_id.in_(branch_ids))
    if search:
        pattern = f"%{search.lower()}%"
        query = query.filter(
            or_(
                func.lower(models.Transaction.account_name).like(pattern),
                func.lower(models.Transaction.detail).like(pattern),
                func.lower(models.Transaction.note).like(pattern),
                func.lower(models.Transaction.category).like(pattern),
            )
        )
    query = query.order_by(models.Transaction.date.asc(), models.Transaction.id.asc())
    if skip:
        query = query.offset(skip)
    if limit is not None:
        query = query.limit(limit)
    return query.all()


def account_ledger(db: Session, account_id: int) -> dict:
    account = get_account(db, account_id)
    if not account:
        return {}
    transactions = (
        db.query(models.Transaction)
        .filter(models.Transaction.account_id == account_id)
        .order_by(models.Transaction.date.asc(), models.Transaction.id.asc())
        .all()
    )
    running_afn = round(account.opening_balance_afn or 0, 2)
    running_usd = round(account.opening_balance_usd or 0, 2)
    rows = []
    for tx in transactions:
        if tx.transaction_type == "cash_in":
            running_afn += _amount(tx.cash_in_afn)
            running_usd += _amount(tx.usd_in)
        else:
            running_afn -= _amount(tx.cash_out_afn)
            running_usd -= _amount(tx.usd_out)
        rows.append(
            {
                "id": tx.id,
                "transaction_no": tx.transaction_no,
                "date": tx.date,
                "account_name": tx.account_name,
                "detail": tx.detail,
                "cash_in_afn": tx.cash_in_afn,
                "cash_out_afn": tx.cash_out_afn,
                "balance": round(running_afn, 2),
                "usd_in": tx.usd_in,
                "usd_out": tx.usd_out,
                "usd_balance": round(running_usd, 2),
                "note": tx.note,
                "transaction_type": tx.transaction_type,
                "exchange_rate": tx.exchange_rate,
                "payment_method": tx.payment_method,
                "category": tx.category,
            }
        )
    return {
        "account": account,
        "opening_balance_afn": account.opening_balance_afn,
        "opening_balance_usd": account.opening_balance_usd,
        "total_cash_in_afn": round(
            sum(_amount(tx.cash_in_afn) for tx in transactions), 2
        ),
        "total_cash_out_afn": round(
            sum(_amount(tx.cash_out_afn) for tx in transactions), 2
        ),
        "total_usd_in": round(sum(_amount(tx.usd_in) for tx in transactions), 2),
        "total_usd_out": round(sum(_amount(tx.usd_out) for tx in transactions), 2),
        "final_balance_afn": round(running_afn, 2),
        "final_balance_usd": round(running_usd, 2),
        "rows": rows,
    }


def backup_payload(db: Session) -> dict:
    db.add(
        models.BackupLog(
            backup_name=f"cashbook-backup-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}",
            backup_type="export",
            note="Full JSON backup exported",
        )
    )
    db.commit()
    return {
        "exported_at": datetime.utcnow(),
        "settings": get_settings(db),
        "accounts": list_accounts(db),
        "employees": db.query(models.Employee)
        .order_by(models.Employee.full_name.asc())
        .all(),
        "transactions": list_transactions(db),
        "salary_payments": db.query(models.SalaryPayment)
        .order_by(models.SalaryPayment.id.asc())
        .all(),
        "salary_history": db.query(models.SalaryHistory)
        .order_by(models.SalaryHistory.id.asc())
        .all(),
        "salary_adjustments": db.query(models.EmployeeSalaryAdjustment)
        .order_by(models.EmployeeSalaryAdjustment.id.asc())
        .all(),
    }


def import_backup(db: Session, payload: dict, replace_all: bool = False) -> dict:
    payload = _backup_root(payload)
    settings_data = _normalize_settings_backup(payload.get("settings") or {})
    account_rows = [
        row
        for row in (
            _normalize_account_backup(item)
            for item in _rows_from(payload, "accounts", "account_list", "parties")
        )
        if row
    ]
    employee_rows = [
        (
            _first_value(item, "id", "employee_id", "employeeId"),
            row,
        )
        for item in _rows_from(payload, "employees", "employee_list", "staff")
        if (row := _normalize_employee_backup(item))
    ]
    transaction_rows = [
        (
            _first_value(item, "id", "transaction_id", "transactionId"),
            row,
        )
        for item in _rows_from(
            payload, "transactions", "records", "entries", "cashbook"
        )
        if (row := _normalize_transaction_backup(item))
    ]
    salary_history_rows = _rows_from(payload, "salary_history", "salaryHistory")
    salary_payment_rows = _rows_from(payload, "salary_payments", "salaryPayments")
    salary_adjustment_rows = _rows_from(
        payload, "salary_adjustments", "salaryAdjustments"
    )

    if replace_all:
        db.query(models.EmployeeSalaryAdjustment).delete()
        db.query(models.SalaryPayment).delete()
        db.query(models.SalaryHistory).delete()
        db.query(models.Transaction).delete()
        db.query(models.Employee).delete()
        db.query(models.Account).delete()
        db.query(models.Setting).delete()
        db.commit()
        db.expunge_all()
        _ensure_settings(db)
    if settings_data:
        update_settings(db, schemas.SettingUpdate(**settings_data))
    imported_accounts = 0
    imported_employees = 0
    imported_transactions = 0
    imported_salary_payments = 0
    imported_salary_history = 0
    imported_salary_adjustments = 0
    for account_data in account_rows:
        account = get_account_by_name(db, account_data["name"])
        if account:
            update_account(db, account, schemas.AccountUpdate(**account_data))
        else:
            create_account(db, schemas.AccountCreate(**account_data))
        imported_accounts += 1
    employee_id_map = {}
    if employee_rows:
        from .payroll import create_employee

        for original_id, employee_data in employee_rows:
            existing_employee = (
                db.query(models.Employee)
                .filter(
                    func.lower(models.Employee.full_name)
                    == employee_data["full_name"].lower()
                )
                .first()
            )
            employee = existing_employee or create_employee(
                db, schemas.EmployeeCreate(**employee_data)
            )
            _map_imported(employee_id_map, original_id, employee)
            imported_employees += 1
    transaction_id_map = {}
    for original_transaction_id, tx_data in transaction_rows:
        if not any(
            (
                _amount(tx_data.get("cash_in_afn")),
                _amount(tx_data.get("cash_out_afn")),
                _amount(tx_data.get("usd_in")),
                _amount(tx_data.get("usd_out")),
            )
        ):
            continue
        existing = (
            get_transaction(db, tx_data["id"])
            if isinstance(tx_data.get("id"), int)
            else None
        )
        if existing:
            _map_imported(transaction_id_map, original_transaction_id, existing)
            continue
        natural_duplicate = (
            db.query(models.Transaction)
            .filter(
                models.Transaction.date == tx_data["date"],
                func.lower(models.Transaction.account_name)
                == tx_data["account_name"].lower(),
                models.Transaction.detail == tx_data["detail"],
                models.Transaction.transaction_type
                == tx_data.get("transaction_type", tx_data.get("type")),
                models.Transaction.cash_in_afn
                == _amount(tx_data.get("cash_in_afn", 0)),
                models.Transaction.cash_out_afn
                == _amount(tx_data.get("cash_out_afn", 0)),
                models.Transaction.usd_in == _amount(tx_data.get("usd_in", 0)),
                models.Transaction.usd_out == _amount(tx_data.get("usd_out", 0)),
                models.Transaction.exchange_rate
                == _amount(tx_data.get("exchange_rate", 0)),
                models.Transaction.note == tx_data.get("note", ""),
            )
            .first()
        )
        if natural_duplicate:
            _map_imported(
                transaction_id_map, original_transaction_id, natural_duplicate
            )
            continue
        account = get_account_by_name(db, tx_data["account_name"])
        account_id = account.id if account else tx_data.get("account_id")
        employee = _get_imported(employee_id_map, tx_data.get("employee_id"))
        created_transaction = create_transaction(
            db,
            schemas.TransactionCreate(
                date=tx_data["date"],
                account_id=account_id,
                employee_id=employee.id if employee else None,
                salary_month=tx_data.get("salary_month"),
                payroll_kind=tx_data.get("payroll_kind"),
                account_name=tx_data["account_name"],
                detail=tx_data["detail"],
                transaction_type=tx_data.get("transaction_type", tx_data.get("type")),
                cash_in_afn=tx_data.get("cash_in_afn", 0),
                cash_out_afn=tx_data.get("cash_out_afn", 0),
                usd_in=tx_data.get("usd_in", 0),
                usd_out=tx_data.get("usd_out", 0),
                exchange_rate=tx_data.get("exchange_rate", 0),
                converted_afn=tx_data.get("converted_afn", 0),
                payment_method=tx_data.get("payment_method", "cash"),
                category=tx_data.get("category", "other"),
                note=tx_data.get("note", ""),
            ),
        )
        _map_imported(transaction_id_map, original_transaction_id, created_transaction)
        imported_transactions += 1

    for history_data in salary_history_rows:
        if not isinstance(history_data, dict):
            continue
        employee = _get_imported(
            employee_id_map, _first_value(history_data, "employee_id", "employeeId")
        )
        if not employee:
            continue
        effective_date = _date_value(
            _first_value(history_data, "effective_date", "effectiveDate", "date")
        )
        if not effective_date:
            continue
        duplicate = (
            db.query(models.SalaryHistory)
            .filter(
                models.SalaryHistory.employee_id == employee.id,
                models.SalaryHistory.effective_date == effective_date,
                models.SalaryHistory.new_salary
                == _amount(
                    _first_value(history_data, "new_salary", "newSalary", "salary")
                ),
            )
            .first()
        )
        if duplicate:
            continue
        db.add(
            models.SalaryHistory(
                employee_id=employee.id,
                old_salary=_amount(
                    _first_value(history_data, "old_salary", "oldSalary")
                ),
                new_salary=_amount(
                    _first_value(history_data, "new_salary", "newSalary", "salary")
                ),
                old_currency=str(
                    _first_value(
                        history_data, "old_currency", "oldCurrency", default="AFN"
                    )
                    or "AFN"
                ).upper(),
                new_currency=str(
                    _first_value(
                        history_data,
                        "new_currency",
                        "newCurrency",
                        "currency",
                        default="AFN",
                    )
                    or "AFN"
                ).upper(),
                effective_date=effective_date,
                changed_at=_datetime_value(
                    _first_value(history_data, "changed_at", "changedAt")
                )
                or datetime.utcnow(),
                changed_by=_normalize_text(
                    _first_value(history_data, "changed_by", "changedBy")
                )
                or "Administrator",
                reason=_normalize_text(_first_value(history_data, "reason"))
                or "Imported backup",
                notes=_normalize_text(_first_value(history_data, "notes", "note")),
            )
        )
        imported_salary_history += 1

    for payment_data in salary_payment_rows:
        if not isinstance(payment_data, dict):
            continue
        employee = _get_imported(
            employee_id_map, _first_value(payment_data, "employee_id", "employeeId")
        )
        transaction = _get_imported(
            transaction_id_map,
            _first_value(
                payment_data,
                "cashbook_entry_id",
                "cashbookEntryId",
                "transaction_id",
                "transactionId",
            ),
        )
        if not employee or not transaction:
            continue
        duplicate = (
            db.query(models.SalaryPayment)
            .filter(
                models.SalaryPayment.cashbook_entry_id == transaction.id,
            )
            .first()
        )
        if duplicate:
            continue
        payment_date = _date_value(
            _first_value(payment_data, "payment_date", "paymentDate", "date")
        )
        if not payment_date:
            continue
        db.add(
            models.SalaryPayment(
                employee_id=employee.id,
                month=int(
                    _amount(_first_value(payment_data, "month")) or payment_date.month
                ),
                year=int(
                    _amount(_first_value(payment_data, "year")) or payment_date.year
                ),
                amount=_amount(
                    _first_value(payment_data, "amount", "paid_amount", "paidAmount")
                ),
                payment_date=payment_date,
                payment_method=_choice(
                    _first_value(
                        payment_data, "payment_method", "paymentMethod", "method"
                    ),
                    {"cash", "bank", "hawala", "other"},
                    "cash",
                    {"transfer": "bank", "card": "bank"},
                ),
                notes=_normalize_text(_first_value(payment_data, "notes", "note")),
                previous_carry_forward_balance=_amount(
                    _first_value(
                        payment_data,
                        "previous_carry_forward_balance",
                        "previousCarryForwardBalance",
                    )
                ),
                total_payable_salary=_amount(
                    _first_value(
                        payment_data, "total_payable_salary", "totalPayableSalary"
                    )
                ),
                carry_forward_balance=_amount(
                    _first_value(
                        payment_data, "carry_forward_balance", "carryForwardBalance"
                    )
                ),
                cashbook_entry_id=transaction.id,
            )
        )
        imported_salary_payments += 1

    for adj_data in salary_adjustment_rows:
        if not isinstance(adj_data, dict):
            continue
        employee = _get_imported(
            employee_id_map, _first_value(adj_data, "employee_id", "employeeId")
        )
        if not employee:
            continue
        adj_date = _date_value(
            _first_value(adj_data, "date", "adjustment_date", "adjustmentDate")
        )
        if not adj_date:
            continue
        period_str = str(
            _first_value(adj_data, "period")
            or f"{adj_date.year:04d}-{adj_date.month:02d}"
        )
        amount_val = _amount(_first_value(adj_data, "amount"))
        if not amount_val:
            continue
        duplicate = (
            db.query(models.EmployeeSalaryAdjustment)
            .filter(
                models.EmployeeSalaryAdjustment.employee_id == employee.id,
                models.EmployeeSalaryAdjustment.date == adj_date,
                models.EmployeeSalaryAdjustment.amount == amount_val,
                models.EmployeeSalaryAdjustment.adjustment_type
                == str(
                    _first_value(
                        adj_data,
                        "adjustment_type",
                        "adjustmentType",
                        "type",
                        default="adjustment",
                    )
                ),
            )
            .first()
        )
        if duplicate:
            continue
        db.add(
            models.EmployeeSalaryAdjustment(
                employee_id=employee.id,
                date=adj_date,
                period=period_str,
                amount=amount_val,
                currency=str(
                    _first_value(adj_data, "currency", default="AFN") or "AFN"
                ).upper(),
                adjustment_type=str(
                    _first_value(
                        adj_data,
                        "adjustment_type",
                        "adjustmentType",
                        "type",
                        default="adjustment",
                    )
                ),
                reason=_normalize_text(_first_value(adj_data, "reason"))
                or "Imported adjustment",
                notes=_normalize_text(_first_value(adj_data, "notes", "note")),
                created_by=_normalize_text(
                    _first_value(adj_data, "created_by", "createdBy")
                )
                or "Administrator",
                created_at=_datetime_value(
                    _first_value(adj_data, "created_at", "createdAt")
                )
                or datetime.utcnow(),
            )
        )
        imported_salary_adjustments += 1

    db.add(
        models.BackupLog(
            backup_name=f"cashbook-restore-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}",
            backup_type="restore",
            note=f"Imported {imported_accounts} accounts, {imported_employees} employees, and {imported_transactions} transactions",
        )
    )
    db.commit()
    return {
        "imported_accounts": imported_accounts,
        "imported_employees": imported_employees,
        "imported_transactions": imported_transactions,
        "imported_salary_payments": imported_salary_payments,
        "imported_salary_history": imported_salary_history,
        "imported_salary_adjustments": imported_salary_adjustments,
    }


def clear_all(db: Session) -> dict:
    transaction_count = db.query(models.Transaction).count()
    employee_count = db.query(models.Employee).count()
    account_count = db.query(models.Account).count()
    db.query(models.EmployeeSalaryAdjustment).delete()
    db.query(models.SalaryPayment).delete()
    db.query(models.SalaryHistory).delete()
    db.query(models.Transaction).delete()
    db.query(models.Employee).delete()
    db.query(models.Account).delete()
    db.query(models.Setting).delete()
    db.commit()

    db.query(models.Employee).delete()
    db.query(models.Account).delete()
    db.query(models.Setting).delete()
    db.commit()
    db.expunge_all()
    _ensure_settings(db)
    db.add(
        models.BackupLog(
            backup_name=f"cashbook-clear-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}",
            backup_type="clear",
            note=f"Cleared {account_count} accounts, {employee_count} employees, and {transaction_count} transactions",
        )
    )
    db.commit()
    return {
        "ok": True,
        "deleted_accounts": account_count,
        "deleted_employees": employee_count,
        "deleted_transactions": transaction_count,
    }


def dashboard_summary(db: Session, branch_id: int | None = None) -> dict:
    today = date.today()
    month_start = today.replace(day=1)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)

    base_q = db.query(models.Transaction)
    if branch_id:
        base_q = base_q.filter(models.Transaction.branch_id == branch_id)

    today_q = base_q.filter(models.Transaction.date == today)
    month_q = base_q.filter(
        models.Transaction.date >= month_start, models.Transaction.date < next_month
    )

    def get_sum(q, col):
        return _amount(q.with_entities(func.sum(col)).scalar())

    afn_in_today = get_sum(today_q, models.Transaction.cash_in_afn)
    afn_out_today = get_sum(today_q, models.Transaction.cash_out_afn)
    afn_in_month = get_sum(month_q, models.Transaction.cash_in_afn)
    afn_out_month = get_sum(month_q, models.Transaction.cash_out_afn)
    afn_balance = get_sum(base_q, models.Transaction.cash_in_afn) - get_sum(
        base_q, models.Transaction.cash_out_afn
    )

    usd_in_today = get_sum(today_q, models.Transaction.usd_in)
    usd_out_today = get_sum(today_q, models.Transaction.usd_out)
    usd_in_month = get_sum(month_q, models.Transaction.usd_in)
    usd_out_month = get_sum(month_q, models.Transaction.usd_out)
    usd_balance = get_sum(base_q, models.Transaction.usd_in) - get_sum(
        base_q, models.Transaction.usd_out
    )

    toman_balance = 0.0
    toman_in_today = 0.0
    toman_out_today = 0.0
    toman_in_month = 0.0
    toman_out_month = 0.0

    entries_today = today_q.count()
    entries_month = month_q.count()

    active_accounts = db.query(models.Account).count()
    # Handle both boolean and integer is_active representation
    active_employees = (
        db.query(models.Employee)
        .filter(or_(models.Employee.is_active == True, models.Employee.is_active == 1))
        .count()
    )

    recent_transactions = base_q.order_by(models.Transaction.id.desc()).limit(10).all()

    thirty_days_ago = today - timedelta(days=30)
    cf_rows = base_q.filter(models.Transaction.date >= thirty_days_ago).all()
    cf_map = {}
    for r in cf_rows:
        d = r.date.isoformat()
        if d not in cf_map:
            cf_map[d] = {
                "date": d,
                "in_afn": 0,
                "out_afn": 0,
                "in_usd": 0,
                "out_usd": 0,
            }
        cf_map[d]["in_afn"] += _amount(r.cash_in_afn)
        cf_map[d]["out_afn"] += _amount(r.cash_out_afn)
        cf_map[d]["in_usd"] += _amount(r.usd_in)
        cf_map[d]["out_usd"] += _amount(r.usd_out)

    cash_flow = sorted(list(cf_map.values()), key=lambda x: x["date"])

    return {
        "period": {"start": month_start.isoformat(), "end": today.isoformat()},
        "branch": {
            "id": branch_id,
            "name": "Consolidated" if not branch_id else str(branch_id),
        },
        "currencies": {
            "AFN": {
                "balance": round(afn_balance, 2),
                "cash_in_today": round(afn_in_today, 2),
                "cash_out_today": round(afn_out_today, 2),
                "cash_in_month": round(afn_in_month, 2),
                "cash_out_month": round(afn_out_month, 2),
            },
            "USD": {
                "balance": round(usd_balance, 2),
                "cash_in_today": round(usd_in_today, 2),
                "cash_out_today": round(usd_out_today, 2),
                "cash_in_month": round(usd_in_month, 2),
                "cash_out_month": round(usd_out_month, 2),
            },
            "TOMAN": {
                "balance": round(toman_balance, 2),
                "cash_in_today": round(toman_in_today, 2),
                "cash_out_today": round(toman_out_today, 2),
                "cash_in_month": round(toman_in_month, 2),
                "cash_out_month": round(toman_out_month, 2),
            },
        },
        "totals": {
            "entries_today": entries_today,
            "entries_month": entries_month,
            "active_accounts": active_accounts,
            "active_employees": active_employees,
        },
        "cash_flow": cash_flow,
        "recent_transactions": recent_transactions,
        "account_balances": [],
        "system_status": {"status": "operational"},
    }
